"""
导出服务（T5-1）：订单报表 / 调度结果导出

统一输出 bytes，由 api/export.py 包装为下载响应。
- CSV：UTF-8 带 BOM 前缀（Excel 打开中文不乱码）
- XLSX：openpyxl 内存 Workbook → BytesIO
"""
import csv
import io
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models.global_schedule import GlobalSchedule
from models.node import Node
from models.order import Order
from models.package import Package

# 订单状态 → 中文文案
ORDER_STATUS_TEXT = {
    "unassigned": "待分配",
    "assigned": "已分配",
    "in_transit": "运输中",
    "signed": "已签收",
    "exception": "异常",
    "closed": "已关闭",
}

# 订单报表列定义
ORDER_HEADERS = [
    "订单编号", "目的地节点编码", "目的地节点名称", "时效要求",
    "状态", "货物数量", "创建时间", "更新时间",
]

# 调度结果列定义（每个包裹一行，方案汇总列重复填充）
SCHEDULE_HEADERS = [
    "方案编号", "算法", "总距离(km)", "总耗时(h)", "货物总数",
    "评分", "状态", "版本", "创建时间",
    "包裹编号", "起点节点编码", "终点节点编码", "重量(kg)", "体积(m³)",
    "包裹状态", "货物明细",
]

SCHEDULE_STATUS_TEXT = {"active": "生效", "draft": "草稿", "superseded": "已替换"}


def _format_dt(dt) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""


def _rows_to_csv(headers: List[str], rows: List[Dict[str, Any]]) -> bytes:
    """生成带 BOM 的 UTF-8 CSV，Excel 直接打开中文不乱码"""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([row.get(h, "") for h in headers])
    return ("﻿" + buf.getvalue()).encode("utf-8")


def _rows_to_xlsx(headers: List[str], rows: List[Dict[str, Any]]) -> bytes:
    """生成 XLSX（内存写入，不落盘）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导出数据"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    # 自适应列宽（粗设，中文按双倍宽估算）
    for col_idx, header in enumerate(headers, start=1):
        width = max(len(str(header)) * 2, 12)
        for row in rows:
            v = row.get(header, "")
            if isinstance(v, (int, float)):
                continue
            width = max(width, len(str(v)) * 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 60)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def serialize_rows(format: str, headers: List[str], rows: List[Dict[str, Any]]) -> bytes:
    """按 format（csv/xlsx）序列化表格为字节"""
    if format == "csv":
        return _rows_to_csv(headers, rows)
    return _rows_to_xlsx(headers, rows)


# ── 订单报表 ─────────────────────────────────────────────

def _build_order_rows(db: Session) -> List[Dict[str, Any]]:
    """查询全部订单并构造报表行"""
    orders = db.query(Order).order_by(Order.created_at.desc()).all()
    rows = []
    for order in orders:
        dest = db.query(Node).filter(Node.id == order.destination_node_id).first()
        rows.append({
            "订单编号": order.order_code,
            "目的地节点编码": dest.node_code if dest else "",
            "目的地节点名称": dest.name if dest else "",
            "时效要求": order.time_window,
            "状态": ORDER_STATUS_TEXT.get(order.status, order.status),
            "货物数量": len(order.goods) if order.goods else 0,
            "创建时间": _format_dt(order.created_at),
            "更新时间": _format_dt(order.updated_at),
        })
    return rows


def export_orders(format: str, db: Session) -> bytes:
    """导出完整订单表"""
    return serialize_rows(format, ORDER_HEADERS, _build_order_rows(db))


# ── 调度结果 ─────────────────────────────────────────────

def _build_schedule_rows(gs: GlobalSchedule, db: Session) -> List[Dict[str, Any]]:
    """构造调度方案包裹明细行（方案汇总列重复填充）"""
    packages = (
        db.query(Package)
        .filter(Package.schedule_id == gs.id)
        .order_by(Package.package_code)
        .all()
    )

    rows = []
    if not packages:
        # 无包裹时仍输出方案汇总行，便于报表对齐
        rows.append(_schedule_summary_row(gs, None, None))
        return rows

    for pkg in packages:
        from_node = db.query(Node).filter(Node.id == pkg.from_node_id).first()
        to_node = db.query(Node).filter(Node.id == pkg.to_node_id).first()
        goods_desc = "、".join(
            item.get("goods_code", "") for item in (pkg.goods_items or [])
        )
        rows.append(_schedule_summary_row(
            gs,
            pkg,
            {"from": from_node, "to": to_node, "goods_desc": goods_desc},
        ))
    return rows


def _schedule_summary_row(gs: GlobalSchedule, pkg, pkg_ctx) -> Dict[str, Any]:
    row = {
        "方案编号": gs.schedule_code,
        "算法": gs.algorithm_type,
        "总距离(km)": float(gs.total_distance),
        "总耗时(h)": float(gs.total_time),
        "货物总数": gs.total_goods,
        "评分": float(gs.score),
        "状态": SCHEDULE_STATUS_TEXT.get(gs.status, gs.status),
        "版本": gs.version,
        "创建时间": _format_dt(gs.created_at),
    }
    if pkg is None:
        row.update({
            "包裹编号": "", "起点节点编码": "", "终点节点编码": "",
            "重量(kg)": "", "体积(m³)": "", "包裹状态": "", "货物明细": "",
        })
    else:
        row.update({
            "包裹编号": pkg.package_code,
            "起点节点编码": pkg_ctx["from"].node_code if pkg_ctx["from"] else "",
            "终点节点编码": pkg_ctx["to"].node_code if pkg_ctx["to"] else "",
            "重量(kg)": float(pkg.weight),
            "体积(m³)": float(pkg.volume),
            "包裹状态": pkg.status,
            "货物明细": pkg_ctx["goods_desc"],
        })
    return row


def export_schedule(format: str, schedule_code: str, db: Session) -> bytes:
    """导出指定调度方案结果

    Raises:
        ValueError: 方案不存在
    """
    gs = (
        db.query(GlobalSchedule)
        .filter(GlobalSchedule.schedule_code == schedule_code)
        .first()
    )
    if not gs:
        raise ValueError(f"调度方案不存在: {schedule_code}")
    return serialize_rows(format, SCHEDULE_HEADERS, _build_schedule_rows(gs, db))
