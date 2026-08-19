from typing import List, Dict, Any
from sqlalchemy.orm import Session, joinedload
from fastapi import UploadFile
from datetime import datetime
import openpyxl
import tempfile
import os
import random
import json

from models.order import Order
from models.goods import Goods
from models.node import Node
from models.sorting_center import SortingCenter
from schemas.order import OrderCreate, OrderUpdate, OrderResponse, OrderImportResponse
from core.error_codes import (CODE_SUCCESS, CODE_PARAM_ERROR, CODE_INTERNAL_ERROR,
                             CODE_ORDER_NOT_FOUND, CODE_ORDER_STATUS_NOT_ALLOWED,
                             CODE_NODE_NOT_FOUND, CODE_ORDER_IMPORT_FAILED)
from core.order_status import (
    ORDER_CLOSABLE_STATUSES,
    ORDER_MUTABLE_STATUSES,
    ORDER_UNASSIGNED,
    is_known_order_status,
    unknown_order_status_message,
)


class OrderService:
    """订单服务"""

    @staticmethod
    async def create_order(order_create: OrderCreate, db: Session) -> Dict[str, Any]:
        """创建订单"""
        try:
            # 1. 校验destination_node_code是否存在
            dest_node = db.query(Node).filter(Node.node_code == order_create.destination_node_code).first()
            if not dest_node:
                return {"code": CODE_NODE_NOT_FOUND, "message": "目的地节点不存在", "data": None}
            
            # 1.5 校验目的地节点必须是0级分拣中心
            if dest_node.node_type != "sorting_center":
                return {"code": CODE_PARAM_ERROR, "message": "目的地必须是分拣中心", "data": None}
            
            # 查询sorting_center表校验level
            sorting_center = db.query(SortingCenter).filter(SortingCenter.node_id == dest_node.id).first()
            if not sorting_center or sorting_center.level != 0:
                return {"code": CODE_PARAM_ERROR, "message": "目的地必须是0级分拣中心", "data": None}

            # 1.6 确定存储中心（货物起点）
            storage_center_code = order_create.storage_center_code
            if storage_center_code:
                # 指定了存储中心，校验存在且类型正确
                sc_node = db.query(Node).filter(
                    Node.node_code == storage_center_code,
                    Node.node_type == "storage_center"
                ).first()
                if not sc_node:
                    return {"code": CODE_PARAM_ERROR, "message": "指定的存储中心不存在", "data": None}
            else:
                # 未指定则随机分配一个存储中心
                sc_nodes = db.query(Node).filter(Node.node_type == "storage_center").all()
                if not sc_nodes:
                    return {"code": CODE_INTERNAL_ERROR, "message": "系统中无可用存储中心，请先初始化演示数据", "data": None}
                sc_node = random.choice(sc_nodes)

            # 2. 生成order_code
            import time
            order_code = f"O{int(time.time() * 1000)}"

            # 3. 创建Order记录
            order = Order(
                order_code=order_code,
                destination_node_id=dest_node.id,
                time_window=order_create.time_window,
                status=ORDER_UNASSIGNED
            )
            db.add(order)
            db.flush()  # 获取order.id

            # 4. 创建Goods记录（node_id 设为存储中心，即货物起始位置）
            for idx, goods_item in enumerate(order_create.goods):
                goods_code = f"G{int(time.time() * 1000)}_{idx}"
                goods = Goods(
                    goods_code=goods_code,
                    order_id=order.id,
                    goods_name=goods_item.goods_name,
                    goods_type=goods_item.goods_type,
                    weight=goods_item.weight,
                    volume=goods_item.volume,
                    node_id=sc_node.id,
                    status="pending_pack"
                )
                db.add(goods)

            db.commit()

            # 5. 返回响应
            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code,
                    "destination_node_name": dest_node.name,
                    "storage_center_code": sc_node.node_code,
                    "storage_center_name": sc_node.name,
                    "time_window": order.time_window,
                    "status": order.status,
                    "goods_count": len(order_create.goods),
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat()
                }
            }
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"创建订单失败: {str(e)}", "data": None}

    @staticmethod
    async def get_orders(page: int, page_size: int, status: str = None, db: Session = None) -> Dict[str, Any]:
        """获取订单列表"""
        try:
            # 1. 构建查询（使用joinedload预加载关联对象，减少N+1查询）
            query = db.query(Order).options(
                joinedload(Order.destination_node),
                joinedload(Order.goods)
            )
            if status:
                if not is_known_order_status(status):
                    return {
                        "code": CODE_PARAM_ERROR,
                        "message": unknown_order_status_message(status),
                        "data": None,
                    }
                query = query.filter(Order.status == status)

            # 2. 分页
            total = query.count()
            orders = query.offset((page - 1) * page_size).limit(page_size).all()

            # 3. 构建响应
            items = []
            for order in orders:
                dest_node = order.destination_node
                goods_count = len(order.goods) if order.goods else 0
                items.append({
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code if dest_node else "",
                    "destination_node_name": dest_node.name if dest_node else "",
                    "time_window": order.time_window,
                    "status": order.status,
                    "goods_count": goods_count,
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat()
                })

            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "items": items,
                    "total": total,
                    "page": page,
                    "page_size": page_size
                }
            }
        except Exception as e:
            return {"code": CODE_INTERNAL_ERROR, "message": f"获取订单列表失败: {str(e)}", "data": None}

    @staticmethod
    async def get_order(order_code: str, db: Session) -> Dict[str, Any]:
        """获取订单详情"""
        try:
            # 1. 查询Order
            order = db.query(Order).filter(Order.order_code == order_code).first()
            if not order:
                return {"code": CODE_ORDER_NOT_FOUND, "message": "订单不存在", "data": None}

            # 2. 获取目的地节点
            dest_node = db.query(Node).filter(Node.id == order.destination_node_id).first()

            # 3. 获取货物列表
            goods = db.query(Goods).filter(Goods.order_id == order.id).all()
            goods_list = []
            for g in goods:
                goods_list.append({
                    "goods_code": g.goods_code,
                    "goods_name": g.goods_name,
                    "goods_type": g.goods_type,
                    "weight": float(g.weight),
                    "volume": float(g.volume),
                    "status": g.status
                })

            # 4. 返回响应
            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code if dest_node else "",
                    "destination_node_name": dest_node.name if dest_node else "",
                    "time_window": order.time_window,
                    "status": order.status,
                    "goods": goods_list,
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat()
                }
            }
        except Exception as e:
            return {"code": CODE_INTERNAL_ERROR, "message": f"获取订单详情失败: {str(e)}", "data": None}

    @staticmethod
    async def update_order(order_code: str, order_update: OrderUpdate, db: Session) -> Dict[str, Any]:
        """更新订单"""
        try:
            # 1. 查询Order
            order = db.query(Order).filter(Order.order_code == order_code).first()
            if not order:
                return {"code": CODE_ORDER_NOT_FOUND, "message": "订单不存在", "data": None}

            # 2. 校验订单状态（仅 unassigned/assigned 可修改）
            if order.status not in ORDER_MUTABLE_STATUSES:
                return {"code": CODE_ORDER_STATUS_NOT_ALLOWED, "message": "订单状态不允许修改", "data": None}

            # 3. 更新字段
            if order_update.destination_node_code is not None:
                dest_node = db.query(Node).filter(Node.node_code == order_update.destination_node_code).first()
                if not dest_node:
                    return {"code": CODE_NODE_NOT_FOUND, "message": "目的地节点不存在", "data": None}
                order.destination_node_id = dest_node.id

            if order_update.time_window is not None:
                order.time_window = order_update.time_window

            order.updated_at = datetime.now()
            db.commit()

            # 4. 返回响应
            dest_node = db.query(Node).filter(Node.id == order.destination_node_id).first()
            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code if dest_node else "",
                    "destination_node_name": dest_node.name if dest_node else "",
                    "time_window": order.time_window,
                    "status": order.status,
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat()
                }
            }
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"更新订单失败: {str(e)}", "data": None}

    @staticmethod
    async def delete_order(order_code: str, db: Session) -> Dict[str, Any]:
        """删除订单"""
        try:
            # 1. 查询Order
            order = db.query(Order).filter(Order.order_code == order_code).first()
            if not order:
                return {"code": CODE_ORDER_NOT_FOUND, "message": "订单不存在", "data": None}

            # 2. 校验订单状态（仅 unassigned/assigned 可删除）
            if order.status not in ORDER_MUTABLE_STATUSES:
                return {"code": CODE_ORDER_STATUS_NOT_ALLOWED, "message": "订单状态不允许删除", "data": None}

            # 3. 先删除关联的goods（避免NOT NULL约束错误）
            goods_list = db.query(Goods).filter(Goods.order_id == order.id).all()
            for goods in goods_list:
                db.delete(goods)

            # 4. 删除订单
            db.delete(order)
            db.commit()

            return {"code": CODE_SUCCESS, "message": "success", "data": None}
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"删除订单失败: {str(e)}", "data": None}

    @staticmethod
    async def import_orders(file: UploadFile, skip_errors: bool, db: Session,
                            column_mapping: str = None) -> Dict[str, Any]:
        """批量导入订单（T5-1 增强：自定义列映射 + 错误行报告）

        - column_mapping：JSON 字符串，{"文件表头名": "系统字段名"}
        - skip_errors=True：错误行跳过并返回 failed_rows；
          False：存在任一错误行则整体回滚，返回 CODE_ORDER_IMPORT_FAILED
        """
        try:
            # 1. 读取文件
            contents = await file.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(contents)
                tmp_path = tmp.name

            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active
            os.unlink(tmp_path)  # 删除临时文件

            # 2. 解析表头（支持自定义列映射）
            headers = [cell.value for cell in ws[1]]

            # 列映射：文件列名 → 系统字段名；未映射的列保留原名（即默认表头即系统字段名）
            mapping = {}
            if column_mapping:
                try:
                    parsed = json.loads(column_mapping)
                    if isinstance(parsed, dict):
                        mapping = parsed
                except Exception:
                    return {
                        "code": CODE_PARAM_ERROR,
                        "message": "column_mapping 必须是合法 JSON 对象",
                        "data": None,
                    }

            def to_sys_field(header):
                if header is None:
                    return None
                return mapping.get(header, header)

            # 3. 逐行校验并处理
            success_count = 0
            failed_count = 0
            failed_rows = []
            encountered_error = False

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if row is None or all(cell is None for cell in row):
                    continue
                try:
                    # 解析行数据（应用列映射）
                    raw = {
                        headers[i]: row[i]
                        for i in range(min(len(headers), len(row)))
                    }
                    row_data = {to_sys_field(k): v for k, v in raw.items()}
                    destination_node_code = row_data.get("destination_node_code")
                    storage_center_code = row_data.get("storage_center_code")  # 可选列
                    time_window = row_data.get("time_window")
                    goods_name = row_data.get("goods_name")
                    goods_type = row_data.get("goods_type")
                    weight = row_data.get("weight")
                    volume = row_data.get("volume")

                    # 校验必填字段
                    if not all([destination_node_code, time_window, goods_name, goods_type, weight, volume]):
                        raise ValueError("必填字段不能为空")

                    # 校验目的地节点是否存在
                    dest_node = db.query(Node).filter(Node.node_code == destination_node_code).first()
                    if not dest_node:
                        raise ValueError(f"目的地节点不存在: {destination_node_code}")

                    # 确定存储中心
                    if storage_center_code:
                        sc_node = db.query(Node).filter(
                            Node.node_code == storage_center_code,
                            Node.node_type == "storage_center"
                        ).first()
                        if not sc_node:
                            raise ValueError(f"存储中心不存在: {storage_center_code}")
                    else:
                        sc_nodes = db.query(Node).filter(Node.node_type == "storage_center").all()
                        if not sc_nodes:
                            raise ValueError("系统中无可用存储中心")
                        sc_node = random.choice(sc_nodes)

                    # 创建订单
                    import time
                    order_code = f"O{int(time.time() * 1000)}_{row_idx}"
                    order = Order(
                        order_code=order_code,
                        destination_node_id=dest_node.id,
                        time_window=time_window,
                        status=ORDER_UNASSIGNED
                    )
                    db.add(order)
                    db.flush()

                    # 创建货物（node_id 设为存储中心）
                    goods_code = f"G{int(time.time() * 1000)}_{row_idx}"
                    goods = Goods(
                        goods_code=goods_code,
                        order_id=order.id,
                        goods_name=goods_name,
                        goods_type=goods_type,
                        weight=weight,
                        volume=volume,
                        node_id=sc_node.id,
                        status="pending_pack"
                    )
                    db.add(goods)

                    success_count += 1
                except Exception as e:
                    encountered_error = True
                    failed_count += 1
                    failed_rows.append({
                        "row": row_idx,
                        "error": str(e)
                    })
                    if not skip_errors:
                        break  # 停止处理，稍后整体回滚

            # skip_errors=False 且存在错误行 → 整体回滚
            if encountered_error and not skip_errors:
                db.rollback()
                return {
                    "code": CODE_ORDER_IMPORT_FAILED,
                    "message": "导入失败：存在错误行，已整体回滚",
                    "data": {
                        "success_count": 0,
                        "failed_count": failed_count,
                        "failed_rows": failed_rows,
                    },
                }

            db.commit()

            # 4. 返回结果（含错误行报告）
            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "success_count": success_count,
                    "failed_count": failed_count,
                    "failed_rows": failed_rows
                }
            }
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"导入订单失败: {str(e)}", "data": None}

    @staticmethod
    async def close_order(order_code: str, db: Session) -> Dict[str, Any]:
        """关闭订单（T1-1 新增：unassigned/assigned → closed）"""
        try:
            from services.state_machine import transition_order_status

            # 1. 查询Order
            order = db.query(Order).filter(Order.order_code == order_code).first()
            if not order:
                return {"code": CODE_ORDER_NOT_FOUND, "message": "订单不存在", "data": None}

            # 2. 校验状态：仅 unassigned / assigned 可关闭
            if order.status not in ORDER_CLOSABLE_STATUSES:
                return {
                    "code": CODE_ORDER_STATUS_NOT_ALLOWED,
                    "message": f"订单状态 '{order.status}' 不允许关闭（仅 unassigned/assigned 可关闭）",
                    "data": None
                }

            # 3. 执行状态转换
            try:
                transition_order_status(db, order, "closed")
            except ValueError as e:
                return {"code": CODE_ORDER_STATUS_NOT_ALLOWED, "message": str(e), "data": None}

            order.updated_at = datetime.now()
            db.commit()
            dest_node = db.query(Node).filter(Node.id == order.destination_node_id).first()
            return {
                "code": CODE_SUCCESS,
                "message": "订单已关闭",
                "data": {
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code if dest_node else "",
                    "destination_node_name": dest_node.name if dest_node else "",
                    "time_window": order.time_window,
                    "status": order.status,
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat(),
                }
            }
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"关闭订单失败: {str(e)}", "data": None}

