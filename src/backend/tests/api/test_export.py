"""
导出 API 测试（T5-1）

测试端点：
- POST /api/export/orders?format=csv|xlsx - 导出完整订单表
- POST /api/export/schedule?schedule_code=xxx - 导出调度结果
"""
import io

import openpyxl
import pytest
from sqlalchemy.orm import Session

from models.global_schedule import GlobalSchedule
from models.package import Package
from models.node import Node


@pytest.fixture
def auth_headers(client, test_users):
    """认证头（调度员）"""
    response = client.post("/api/auth/login", json={
        "username": "dispatcher",
        "password": "123456",
    })
    token = response.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
def manager_headers(client, test_users):
    """认证头（管理者，无调度权限）"""
    response = client.post("/api/auth/login", json={
        "username": "manager",
        "password": "123456",
    })
    token = response.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _seed_schedule(db_session, test_nodes, schedule_code="GS_EXPORT"):
    """直接落库一个调度方案 + 包裹"""
    gs = GlobalSchedule(
        schedule_code=schedule_code,
        order_codes=["O001"],
        goods_schedules=[{"goods_code": "G001", "order_code": "O001", "path": ["SC001", "SO010"]}],
        total_distance=100.5,
        total_time=3.2,
        total_goods=1,
        score=0.5,
        algorithm_type="traditional",
        status="active",
        version=1,
    )
    db_session.add(gs)
    db_session.flush()

    pkg = Package(
        package_code=f"PKG_{schedule_code}",
        weight=10.0,
        volume=0.5,
        status="pending_dispatch",
        from_node_id=test_nodes["SC001"].id,
        to_node_id=test_nodes["SO010"].id,
        goods_items=[{"goods_code": "G001", "order_code": "O001"}],
        schedule_id=gs.id,
    )
    db_session.add(pkg)
    db_session.commit()
    return gs


@pytest.mark.api
@pytest.mark.phase5
class TestExportOrders:
    def test_export_orders_csv(self, client, auth_headers, test_orders):
        """导出订单 CSV：返回下载文件且含表头与订单数据"""
        response = client.post("/api/export/orders?format=csv", headers=auth_headers)
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "attachment" in response.headers.get("content-disposition", "")
        # 带 BOM 解码
        text = response.content.decode("utf-8-sig")
        assert "订单编号" in text
        assert "O001" in text

    def test_export_orders_xlsx(self, client, auth_headers, test_orders, test_goods):
        """导出订单 XLSX：可被 openpyxl 解析且含数据行"""
        response = client.post("/api/export/orders?format=xlsx", headers=auth_headers)
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "订单编号"  # 表头
        assert len(rows) >= 2            # 表头 + 至少一行数据


    @pytest.mark.parametrize("redis_mode", ["disabled", "write-error"])
    def test_keyed_export_replays_from_database_when_redis_unavailable(
        self,
        client,
        auth_headers,
        test_orders,
        monkeypatch,
        redis_mode,
    ):
        """Redis 关闭或写入失败时，数据库仍保真重放导出响应。"""
        from services import export_service
        from utils import cache

        calls = 0
        original = export_service.export_orders

        def counted(*args, **kwargs):
            nonlocal calls
            calls += 1
            return original(*args, **kwargs)

        if redis_mode == "disabled":
            monkeypatch.setattr(cache, "resolve_redis", lambda: None)
        else:
            class FailingRedis:
                async def setex(self, *_args, **_kwargs):
                    raise RuntimeError("redis unavailable")

            monkeypatch.setattr(cache, "resolve_redis", lambda: FailingRedis())

        monkeypatch.setattr(export_service, "export_orders", counted)
        headers = {
            **auth_headers,
            "X-Idempotency-Key": f"export-redis-{redis_mode}",
        }

        first = client.post(
            "/api/export/orders?format=xlsx",
            headers=headers,
        )
        replay = client.post(
            "/api/export/orders?format=xlsx",
            headers=headers,
        )

        assert first.status_code == 200
        assert replay.status_code == first.status_code
        assert replay.content == first.content
        assert replay.headers["content-type"] == first.headers["content-type"]
        assert (
            replay.headers["content-disposition"]
            == first.headers["content-disposition"]
        )
        assert calls == 1
        assert not any(
            key.startswith("idem:") for key in cache.memory_cache._store
        )

    def test_export_orders_requires_dispatcher(self, client, manager_headers):
        """非调度角色（manager）导出被拒绝"""
        response = client.post("/api/export/orders", headers=manager_headers)
        assert response.status_code == 403

    def test_export_orders_requires_auth(self, client):
        """未登录导出返回 401"""
        response = client.post("/api/export/orders")
        assert response.status_code == 401

    def test_export_orders_invalid_format(self, client, auth_headers):
        """非法格式返回 422"""
        response = client.post("/api/export/orders?format=pdf", headers=auth_headers)
        assert response.status_code == 422


@pytest.mark.api
@pytest.mark.phase5
class TestExportSchedule:
    def test_export_schedule_xlsx(self, client, auth_headers, db_session, test_nodes):
        """导出调度结果 XLSX：含方案与包裹明细"""
        _seed_schedule(db_session, test_nodes)
        response = client.post(
            "/api/export/schedule?schedule_code=GS_EXPORT&format=xlsx",
            headers=auth_headers,
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [c for c in rows[0]]
        assert "方案编号" in headers
        assert "包裹编号" in headers
        assert "GS_EXPORT" in [r[headers.index("方案编号")] for r in rows[1:]]
        assert "PKG_GS_EXPORT" in [r[headers.index("包裹编号")] for r in rows[1:]]

    def test_export_schedule_csv(self, client, auth_headers, db_session, test_nodes):
        """导出调度结果 CSV"""
        _seed_schedule(db_session, test_nodes)
        response = client.post(
            "/api/export/schedule?schedule_code=GS_EXPORT&format=csv",
            headers=auth_headers,
        )
        assert response.status_code == 200
        text = response.content.decode("utf-8-sig")
        assert "方案编号" in text
        assert "GS_EXPORT" in text

    def test_export_schedule_not_found(self, client, auth_headers):
        """不存在的调度方案返回统一 404 envelope，且不回显方案编号"""
        response = client.post(
            "/api/export/schedule?schedule_code=GS_NOT_EXIST",
            headers=auth_headers,
        )
        assert response.status_code == 404
        body = response.json()
        assert set(body) == {"code", "message", "data", "meta"}
        assert body["code"] == 40400
        assert body["data"] is None
        assert body["message"] == "资源不存在"
        assert "detail" not in body
        assert "GS_NOT_EXIST" not in response.text

    def test_export_schedule_missing_code(self, client, auth_headers):
        """缺少 schedule_code 返回 422"""
        response = client.post("/api/export/schedule", headers=auth_headers)
        assert response.status_code == 422
